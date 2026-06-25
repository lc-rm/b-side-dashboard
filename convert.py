#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ビイサイド（福祉事業部）Indeed広告 & 求人別データ → 非PII JSON 変換器
出力スキーマは AppsScript.gs と共通。data.json は jobcora-dashboard の「広告コスト」「求人別」タブが fetch する。

入力（手元のExcel書き出し）:
  1) ★ビイサイドプランニング様：Indeed結果.xlsx   …「indeedﾃﾞｰﾀ(YYYY年度)」= 月別広告コスト
  2) 【求人別データ2026年度】福祉事業部.xlsx        … 月別シート「YYYY.M」= 求人別Indeed実績

方針: 氏名・電話・メール・正確な年齢などPIIは一切出力しない（GitHub Pages公開のため）。
      ここで扱うのは広告指標と求人(掲載)実績のみ＝非PII。
"""
import json, sys, datetime, re, os
import openpyxl

HOME = os.path.expanduser("~")
INDEED_SRC = sys.argv[1] if len(sys.argv) > 1 else f"{HOME}/Downloads/★ビイサイドプランニング様：Indeed結果.xlsx"
JOBS_SRC   = sys.argv[2] if len(sys.argv) > 2 else f"{HOME}/Downloads/【求人別データ2026年度】福祉事業部.xlsx"
OUT        = sys.argv[3] if len(sys.argv) > 3 else os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(re.sub(r"[^0-9.\-]", "", str(v)))
    except Exception:
        return None


def s(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def yen(v):
    """'¥61' / '5,000' / 61.0 → float"""
    n = num(v)
    return n if n is not None else 0.0


def period_to_ym(v):
    m = re.search(r"(\d{4})/(\d{1,2})", str(v) if v is not None else "")
    return m.group(1) + m.group(2).zfill(2) if m else None


def sheet_ym(name):
    """'2026.6' → '202606'"""
    m = re.match(r"(\d{4})\.(\d{1,2})$", name)
    return m.group(1) + m.group(2).zfill(2) if m else None


# ============================================================ 1) 広告コスト（月別 / indeedﾃﾞｰﾀ）
COST_COLS = ("imp", "click", "CTR", "応募開始数", "CV", "応募完了率", "CVR", "CPC", "CPA", "COST", "予算残")
wb_i = openpyxl.load_workbook(INDEED_SRC, read_only=True, data_only=True)
ad_cost = []
seen_cost = set()
cost_sheets = [n for n in wb_i.sheetnames if n.startswith("indeedﾃﾞｰﾀ")]
for name in cost_sheets:
    rows = list(wb_i[name].iter_rows(values_only=True))
    i = 0
    while i < len(rows):
        r = rows[i]
        if r and s(r[0]) == "indeed応募" and s(r[1]) == "予算":
            idx = {c: ci for ci, c in enumerate(r) if c in COST_COLS}
            tot = rows[i + 1] if i + 1 < len(rows) else None
            ym = None
            camps = []
            j = i + 2
            while j < len(rows) and not (rows[j] and s(rows[j][0]) == "indeed応募"):
                cr = rows[j]
                nm = s(cr[0]) if cr else ""
                if nm and nm != "合計":
                    p = period_to_ym(cr[2])
                    if p and not ym:
                        ym = p
                    g = lambda k, row=cr: (num(row[idx[k]]) if k in idx else None)
                    camps.append({
                        "type": nm, "period": s(cr[2]),
                        "budget": num(cr[1]) or 0, "cost": g("COST") or 0,
                        "imp": int(g("imp") or 0), "click": int(g("click") or 0), "ctr": g("CTR"),
                        "applyStart": int(g("応募開始数") or 0), "cv": int(g("CV") or 0), "cvr": g("CVR"),
                        "cpc": g("CPC"), "cpa": g("CPA"),
                    })
                j += 1
            if tot and s(tot[0]) == "合計" and ym and ym not in seen_cost:
                seen_cost.add(ym)
                g = lambda k: (num(tot[idx[k]]) if k in idx else None)
                ad_cost.append({
                    "ym": ym, "budget": num(tot[1]) or 0, "cost": g("COST") or 0,
                    "imp": int(g("imp") or 0), "click": int(g("click") or 0), "ctr": g("CTR"),
                    "applyStart": int(g("応募開始数") or 0), "cv": int(g("CV") or 0), "cvr": g("CVR"),
                    "cpc": g("CPC"), "cpa": g("CPA"),
                    "budgetLeft": (num(tot[idx["予算残"]]) if "予算残" in idx else None),
                    "campaigns": camps,
                })
            i = j
        else:
            i += 1
ad_cost.sort(key=lambda x: x["ym"])

# ============================================================ 2) 求人別（月別シート YYYY.M）
# 列: 0月日,1CP,2jobKey,3求人,4ステータス,5勤務地,6企業名,7カテゴリー,...,13表示,14クリック,15応募開始,16応募数,...,21クリック単価(CPC)
# ※ 同じ求人が月内に複数回掲載されるので (ym, 求人名) で合算。
#   「応募が発生した求人のみ」を出力して件数とノイズを圧縮（カテゴリー列は不要文字のため除外）。
def cp_norm(v):
    return re.sub(r"^\d{4}年\d{1,2}月[:：]\s*", "", s(v)).strip()

wb_j = openpyxl.load_workbook(JOBS_SRC, read_only=True, data_only=True)
job_sheets = [n for n in wb_j.sheetnames if re.match(r"\d{4}\.\d{1,2}$", n)]
jobs_map = {}  # (ym,title) -> agg
for name in job_sheets:
    ym = sheet_ym(name)
    if not ym:
        continue
    for r in list(wb_j[name].iter_rows(values_only=True))[1:]:
        if not r:
            continue
        title = s(r[3])
        if not title:
            continue
        cp = cp_norm(r[1])
        loc = s(r[5]).split(",")[0].strip()
        imp = int(num(r[13]) or 0)
        click = int(num(r[14]) or 0)
        apply_start = int(num(r[15]) or 0)
        applies = int(num(r[16]) or 0)
        cost = round(click * yen(r[21]))
        key = (ym, title)
        a = jobs_map.get(key)
        if not a:
            a = {"ym": ym, "title": title, "cp": cp, "loc": loc,
                 "imp": 0, "click": 0, "applyStart": 0, "applies": 0, "cost": 0}
            jobs_map[key] = a
        a["imp"] += imp; a["click"] += click; a["applyStart"] += apply_start
        a["applies"] += applies; a["cost"] += cost
        if not a["cp"] and cp:
            a["cp"] = cp
jobs = []
for a in jobs_map.values():
    if a["applies"] <= 0:
        continue  # 応募が発生した求人のみ
    a["cpc"] = round(a["cost"] / a["click"]) if a["click"] else 0
    a["cpa"] = round(a["cost"] / a["applies"]) if a["applies"] else 0
    jobs.append(a)
jobs.sort(key=lambda x: (x["ym"], -x["applies"]))

# ============================================================ meta + 出力
def latest(rows):
    ys = [x["ym"] for x in rows if x.get("ym")]
    return max(ys) if ys else ""

meta = {
    "source": "ビイサイド（福祉事業部）Indeed広告 / 求人別",
    "generatedFrom": "convert.py (local Excel)",
    "fetchedAt": datetime.datetime.now(datetime.timezone.utc).isoformat(),
    "adCostMonths": len(ad_cost),
    "jobRows": len(jobs),
    "latestAdYm": latest(ad_cost),
    "latestJobYm": latest(jobs),
}
out = {"meta": meta, "adCost": ad_cost, "jobs": jobs}
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, separators=(",", ":"))

print("wrote", OUT)
print("adCost months:", len(ad_cost), "| job rows:", len(jobs))
if ad_cost:
    print("ad sample:", json.dumps(ad_cost[-1], ensure_ascii=False)[:300])
if jobs:
    print("job sample:", json.dumps(jobs[0], ensure_ascii=False)[:260])
